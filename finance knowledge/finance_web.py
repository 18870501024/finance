# -*- coding: utf-8 -*-
from pathlib import Path
from datetime import datetime
import time
from flask import Flask, jsonify, render_template_string, abort
import requests
from openpyxl import Workbook, load_workbook
from finance_content import (
    CATEGORIES,
    CARDS,
    CARD_DETAILS,
    CARD_DETAIL_CONTENT,
    CARD_LOOKUP,
    ETFS,
    ETF_CARDS,
    MODELS,
    TEMPLATE,
)
app = Flask(__name__)

for card in ETF_CARDS:
    card['detail_url'] = f"/etfs/{card.get('id', '').strip()}"

ETF_CARD_LOOKUP = {}
for item in ETF_CARDS:
    identifier = (item.get('id') or '').strip()
    if identifier:
        ETF_CARD_LOOKUP[identifier] = item
        ETF_CARD_LOOKUP[identifier.lower()] = item

MODEL_LOOKUP = {}
for model in MODELS:
    identifier = (model.get('id') or '').strip()
    if not identifier:
        continue
    MODEL_LOOKUP[identifier] = model
    MODEL_LOOKUP[identifier.lower()] = model
    model['slug'] = identifier

EASTMONEY_HISTORY_URL = 'https://fundmobapi.eastmoney.com/FundMNewApi/FundMNHisNetList'
EASTMONEY_HEADERS = {
    'User-Agent': 'Mozilla/5.0',
    'Referer': 'https://fund.eastmoney.com/',
}

DATA_DIR = (Path(__file__).resolve().parent / 'data')
DATA_DIR.mkdir(exist_ok=True)
EXCEL_PATH = DATA_DIR / 'etf_monthly.xlsx'
CACHE_MAX_AGE_SECONDS = 60 * 60 * 6
ETF_CACHE = {}
ETF_CACHE_MTIME = 0

ETF_SYMBOL_MAP = {
    '510050': '510050.SS',
    '510300': '510300.SS',
    '510500': '510500.SS',
    '159919': '159919.SZ',
    '512000': '512000.SS',
    '512170': '512170.SS',
    '512800': '512800.SS',
    '515050': '515050.SS',
    '159915': '159915.SZ',
    '516160': '516160.SS',
    '515000': '515000.SS',
    '516970': '516970.SS',
    '510880': '510880.SS',
    '515180': '515180.SS',
    '513130': '513130.SS',
    '159941': '159941.SZ',
    '518880': '518880.SS',
    '512290': '512290.SS',
    '159928': '159928.SZ',
    '510900': '510900.SS',
    '512690': '512690.SS',
    '159949': '159949.SZ',
    '515880': '515880.SS',
    '516770': '516770.SS',
    '159920': '159920.SS',
    '512070': '512070.SS',
    '512960': '512960.SS',
    '513500': '513500.SS',
    '513600': '513600.SS',
    '159967': '159967.SZ',
    '513050': '513050.SS',
    '515680': '515680.SS',
}





def _match_etfs_for_card(card):
    if not card:
        return ETFS

    filters = [flt for flt in (card.get('asset_filters') or []) if flt]
    if not filters:
        return ETFS

    results = []
    for etf in ETFS:
        asset_class = str(etf.get('asset_class') or '')
        if any(flt in asset_class for flt in filters):
            results.append(etf)

    return results or ETFS


def _render_page(page_type: str, **extra_context):
    context = {
        'categories': CATEGORIES,
        'cards': CARDS,
        'models': MODELS,
        'etfs': ETFS,
        'etf_cards': ETF_CARDS,
        'page_type': page_type,
    }
    context.update(extra_context)
    return render_template_string(TEMPLATE, **context)
def _format_series(raw_pairs):
    if not raw_pairs:
        return []

    ordered = sorted(raw_pairs, key=lambda item: item[0])
    if len(ordered) > 35:
        ordered = ordered[-35:]
    base = ordered[0][1] or 1
    if base == 0:
        base = 1

    series = []
    for date_str, value in ordered:
        return_pct = round((value / base - 1) * 100, 2)
        series.append({
            'date': date_str,
            'nav': round(value, 4),
            'return_pct': return_pct,
        })
    return series

def _request_yahoo_series(symbol: str):
    symbol = (symbol or '').upper()
    if not symbol:
        return []

    url = f'https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?interval=1d&range=6mo'
    try:
        response = requests.get(url, timeout=6)
        response.raise_for_status()
        payload = response.json()
    except Exception:
        return []

    result = (payload.get('chart') or {}).get('result')
    if not result:
        return []

    first = result[0]
    timestamps = first.get('timestamp') or []
    quote = (first.get('indicators') or {}).get('quote') or [{}]
    closes = quote[0].get('close') or []

    raw_pairs = []
    for ts, close in zip(timestamps, closes):
        if close in (None, 'null'):
            continue
        try:
            date_str = datetime.fromtimestamp(ts).strftime('%Y-%m-%d')
        except Exception:
            continue
        raw_pairs.append((date_str, float(close)))

    return _format_series(raw_pairs)

def _request_eastmoney_series(code: str):
    code = (code or '').strip()
    if not code:
        return []

    params = {
        'FCODE': code,
        'pageIndex': 1,
        'pageSize': 35,
        'appType': 'ttjj',
        'product': 'EFund',
        'plat': 'Iphone',
        'deviceid': '00000000-0000-0000-0000-000000000000',
        'Version': '1',
        'lf': '1',
        'ctoken': '',
        'userId': '',
        'u': '0',
        'UToken': '',
        'range': '1m',
    }

    try:
        response = requests.get(
            EASTMONEY_HISTORY_URL,
            params=params,
            headers=EASTMONEY_HEADERS,
            timeout=6,
        )
        response.raise_for_status()
        payload = response.json()
    except Exception:
        return []

    items = ((payload.get('Data') or {}).get('LSJZList')) or (payload.get('Datas') or [])
    raw_pairs = []
    for item in items:
        date_str = item.get('FSRQ')
        value_str = item.get('DWJZ')
        if not date_str or value_str in (None, ''):
            continue
        try:
            value = float(value_str)
        except (TypeError, ValueError):
            continue
        raw_pairs.append((date_str, value))

    return _format_series(raw_pairs)


def _write_records_to_excel(records):
    wb = Workbook()
    ws = wb.active
    ws.title = 'etf_history'
    ws.append(['ticker', 'date', 'nav', 'return_pct'])
    for record in records:
        ws.append([
            record['ticker'],
            record['date'],
            record['nav'],
            record['return_pct'],
        ])

    meta = wb.create_sheet('meta')
    meta.append(['generated_at', datetime.utcnow().isoformat(timespec='seconds')])
    wb.save(EXCEL_PATH)


def _load_cache_from_excel():
    if not EXCEL_PATH.exists():
        return {}

    try:
        wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except Exception:
        return {}

    if 'etf_history' not in wb.sheetnames:
        return {}

    sheet = wb['etf_history']
    cache = {}
    for ticker, date_value, nav_value, return_value in sheet.iter_rows(min_row=2, values_only=True):
        if not ticker or not date_value:
            continue
        if isinstance(date_value, datetime):
            date_str = date_value.strftime('%Y-%m-%d')
        else:
            date_str = str(date_value)
        try:
            nav = float(nav_value)
        except (TypeError, ValueError):
            continue
        try:
            pct = float(return_value)
        except (TypeError, ValueError):
            pct = 0.0
        cache.setdefault(ticker.upper(), []).append({
            'date': date_str,
            'nav': round(nav, 4),
            'return_pct': round(pct, 2),
        })

    for series in cache.values():
        series.sort(key=lambda item: item['date'])
    return cache


def _group_records(records):
    grouped = {}
    for record in records:
        grouped.setdefault(record['ticker'], []).append(record)
    for series in grouped.values():
        series.sort(key=lambda item: item['date'])
    return grouped


def _refresh_cache_from_remote():
    aggregated = []
    for etf in ETFS:
        ticker = etf['ticker']
        series = _fetch_remote_etf_series(ticker)
        if not series:
            continue
        for point in series:
            aggregated.append({
                'ticker': ticker,
                'date': point['date'],
                'nav': point['nav'],
                'return_pct': point['return_pct'],
            })

    if not aggregated:
        return False

    _write_records_to_excel(aggregated)
    global ETF_CACHE, ETF_CACHE_MTIME
    ETF_CACHE = _group_records(aggregated)
    ETF_CACHE_MTIME = time.time()
    return True


def ensure_etf_cache(force_refresh=False):
    global ETF_CACHE, ETF_CACHE_MTIME

    if force_refresh:
        if _refresh_cache_from_remote():
            return True

    now = time.time()

    if ETF_CACHE and (now - ETF_CACHE_MTIME) < CACHE_MAX_AGE_SECONDS:
        return True

    if EXCEL_PATH.exists():
        mtime = EXCEL_PATH.stat().st_mtime
        if not ETF_CACHE or ETF_CACHE_MTIME != mtime:
            cache = _load_cache_from_excel()
            if cache:
                ETF_CACHE = cache
                ETF_CACHE_MTIME = mtime
        if ETF_CACHE and (now - mtime) < CACHE_MAX_AGE_SECONDS:
            return True

    if _refresh_cache_from_remote():
        return True

    if ETF_CACHE:
        return True

    if EXCEL_PATH.exists():
        cache = _load_cache_from_excel()
        if cache:
            ETF_CACHE = cache
            ETF_CACHE_MTIME = EXCEL_PATH.stat().st_mtime
            return True

    return False


def fetch_etf_series(ticker: str, force_refresh=False):
    normalized = (ticker or '').upper()
    if not normalized:
        return []

    if not ensure_etf_cache(force_refresh=force_refresh):
        return []

    return ETF_CACHE.get(normalized, [])
def _fetch_remote_etf_series(ticker: str):
    base = (ticker or '').upper()
    if not base:
        return []

    series = _request_eastmoney_series(base)
    if series:
        return series

    candidates = []
    mapped = ETF_SYMBOL_MAP.get(base)
    if mapped:
        candidates.append(mapped.upper())

    for suffix in ['', '.SS', '.SZ']:
        candidate = base if suffix == '' else f"{base}{suffix}"
        if candidate not in candidates:
            candidates.append(candidate)

    for symbol in candidates:
        series = _request_yahoo_series(symbol)
        if series:
            return series

    return []



@app.route('/models/<slug>')
@app.route('/model/<slug>')
def model_detail(slug: str):
    normalized = (slug or '').strip()
    target = MODEL_LOOKUP.get(normalized) or MODEL_LOOKUP.get(normalized.lower())
    if not target:
        abort(404)

    slug_value = target.get('slug') or normalized
    related = [
        item for item in MODELS
        if item.get('slug') != slug_value
    ][:3]

    return _render_page(
        'model_detail',
        model_detail=target,
        related_models=related,
    )


@app.route('/cards/<slug>')
@app.route('/card/<slug>')
def card_detail(slug: str):
    payload = CARD_DETAILS.get(slug) or CARD_DETAILS.get((slug or '').strip())
    if not payload:
        abort(404)

    card = payload['card']
    related = [
        item for item in CARDS
        if item.get('category') == card.get('category') and item.get('slug') != card.get('slug')
    ][:3]

    return _render_page(
        'card_detail',
        card_detail=payload,
        related_cards=related,
    )


@app.route('/etfs/<card_id>')
@app.route('/etf/<card_id>')
def etf_category_detail(card_id: str):
    normalized = (card_id or '').strip()
    card = ETF_CARD_LOOKUP.get(normalized) or ETF_CARD_LOOKUP.get(normalized.lower())
    if not card:
        abort(404)

    matched = _match_etfs_for_card(card)
    primary = matched[0] if matched else None

    return _render_page(
        'etf-detail',
        card=card,
        matched_etfs=matched,
        primary_etf=primary,
    )

@app.route('/api/etf/<ticker>')
def etf_timeseries(ticker: str):
    normalized = (ticker or '').upper()
    if not normalized:
        return jsonify({'dates': [], 'navs': [], 'returns': []}), 400

    if not ensure_etf_cache():
        return jsonify({'dates': [], 'navs': [], 'returns': []}), 502

    series = ETF_CACHE.get(normalized)
    if not series:
        if not ensure_etf_cache(force_refresh=True):
            return jsonify({'dates': [], 'navs': [], 'returns': []}), 502
        series = ETF_CACHE.get(normalized, [])
    if not series:
        return jsonify({'dates': [], 'navs': [], 'returns': []}), 404

    return jsonify({
        'dates': [point['date'] for point in series],
        'navs': [point['nav'] for point in series],
        'returns': [point['return_pct'] for point in series],
    })

@app.route('/')
def index():
    return _render_page('home')


if __name__ == '__main__':
    if not EXCEL_PATH.exists():
        ensure_etf_cache(force_refresh=True)
    else:
        ensure_etf_cache()
    app.run(debug=True)






























































































