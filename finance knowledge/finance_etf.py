# -*- coding: utf-8 -*-
"""ETF data utilities for the finance knowledge web app."""
from pathlib import Path
from datetime import datetime
import time
import requests
from openpyxl import Workbook, load_workbook


EASTMONEY_HISTORY_URL = 'https://fundmobapi.eastmoney.com/FundMNewApi/FundMNHisNetList'
EASTMONEY_HEADERS = {
    'User-Agent': 'Mozilla/5.0',
    'Referer': 'https://fund.eastmoney.com/',
}

DATA_DIR = (Path(__file__).resolve().parent / 'data')
DATA_DIR.mkdir(exist_ok=True)
EXCEL_PATH = DATA_DIR / 'etf_monthly.xlsx'
CACHE_MAX_AGE_SECONDS = 60 * 60 * 6
ETF_CACHE = {}
ETF_CACHE_MTIME = 0

ETF_SYMBOL_MAP = {
    '510050': '510050.SS',
    '510300': '510300.SS',
    '510500': '510500.SS',
    '159919': '159919.SZ',
    '512000': '512000.SS',
    '512170': '512170.SS',
    '512800': '512800.SS',
    '515050': '515050.SS',
    '159915': '159915.SZ',
    '516160': '516160.SS',
    '515000': '515000.SS',
    '516970': '516970.SS',
    '510880': '510880.SS',
    '515180': '515180.SS',
    '513130': '513130.SS',
    '159941': '159941.SZ',
    '518880': '518880.SS',
    '512290': '512290.SS',
    '159928': '159928.SZ',
    '510900': '510900.SS',
    '512690': '512690.SS',
    '159949': '159949.SZ',
    '515880': '515880.SS',
    '516770': '516770.SS',
    '159920': '159920.SS',
    '512070': '512070.SS',
    '512960': '512960.SS',
    '513500': '513500.SS',
    '513600': '513600.SS',
    '159967': '159967.SZ',
    '513050': '513050.SS',
    '515680': '515680.SS',
}


ETF_CARDS = [
    {
        'id': 'all',
        'title_zh': 'ETF 热门全览',
        'title_en': 'Top 20 Universe',
        'description_zh': '覆盖宽基、行业、海外与商品多维度龙头，点击查看完整榜单。',
        'description_en': 'A cross-style mix of 20 widely followed ETFs across markets.',
        'asset_filters': [],
    },
    {
        'id': 'domestic',
        'title_zh': 'A股策略精选',
        'title_en': 'Domestic Focus',
        'description_zh': '优选A股宽基、红利与行业主题ETF，适合打造核心仓位。',
        'description_en': 'Build core China exposure via broad, dividend, and thematic funds.',
        'asset_filters': ['宽基指数', '行业主题', '红利价值', '策略风格', '科技创新', '成长风格'],
    },
    {
        'id': 'global',
        'title_zh': '全球资产补充',
        'title_en': 'Global & Alternatives',
        'description_zh': '配置海外科技与商品ETF，捕捉全球增长与避险机会。',
        'description_en': 'Add offshore tech and commodity sleeves for diversification.',
        'asset_filters': ['海外科技', '海外市场', '商品资产'],
    },
]
TEMPLATE = """<!DOCTYPE html>
<html lang=\"zh-CN\">
<head>
    <meta charset=\"UTF-8\">
    <meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\">
    <title>金融小知识卡片 · Finance Knowledge Cards</title>
    <style>
        :root {
            --accent: #0c7cd5;
            --accent-soft: rgba(12, 124, 213, 0.12);
            --text-dark: #1f2933;
            --text-light: #3e4c59;
            --bg: #f7fafc;
        }

        * {
            box-sizing: border-box;
        }

        body {
            margin: 0;
            font-family: 'Segoe UI', 'Helvetica Neue', Arial, sans-serif;
            background: white;
            color: var(--text-dark);
            line-height: 1.6;
        }

        a {
            color: inherit;
        }

        header {
            border-bottom: 1px solid #e4ebf3;
            background: white;
        }

        .top-bar {
            max-width: 1100px;
            margin: 0 auto;
            padding: 18px 24px;
            display: flex;
            align-items: center;
            justify-content: space-between;
        }

        .brand {
            display: flex;
            align-items: center;
            gap: 12px;
            font-weight: 600;
            font-size: 20px;
            letter-spacing: 0.5px;
            text-transform: capitalize;
        }

        .brand-icon {
            width: 36px;
            height: 36px;
            border-radius: 10px;
            background: var(--accent-soft);
            display: grid;
            place-items: center;
            color: var(--accent);
            font-weight: 700;
        }

        nav {
            display: flex;
            gap: 22px;
            font-size: 15px;
            color: var(--text-light);
        }

        .cta {
            padding: 10px 18px;
            border-radius: 999px;
            background: var(--accent);
            color: white;
            border: none;
            font-size: 14px;
            cursor: pointer;
            transition: transform 0.2s ease, box-shadow 0.2s ease;
        }

        .cta:hover {
            transform: translateY(-2px);
            box-shadow: 0 10px 18px rgba(12, 124, 213, 0.2);
        }

        .hero {
            max-width: 1100px;
            margin: 0 auto;
            padding: 64px 24px 48px;
            display: grid;
            gap: 24px;
        }

        .hero h1 {
            margin: 0;
            font-size: clamp(32px, 5vw, 46px);
            line-height: 1.2;
        }

        .hero .headline-en {
            margin-top: 8px;
            font-size: 20px;
            color: #52606d;
            letter-spacing: 0.5px;
        }

        .hero p {
            max-width: 680px;
            margin: 0;
            color: var(--text-light);
            font-size: 17px;
        }

        .hero .para-en {
            font-size: 15px;
            color: #7b8794;
            margin-top: 6px;
        }

        .filter-bar {
            display: flex;
            flex-wrap: wrap;
            gap: 12px;
            background: var(--bg);
            padding: 12px;
            border-radius: 12px;
            align-items: center;
        }

        .filter-title {
            font-weight: 600;
            font-size: 14px;
            text-transform: uppercase;
            letter-spacing: 1.2px;
            color: var(--text-light);
        }

        .filter-button {
            padding: 9px 16px;
            border-radius: 999px;
            background: white;
            border: 1px solid #d9e2ec;
            cursor: pointer;
            font-size: 13px;
            color: var(--text-light);
            transition: all 0.2s ease;
        }

        .filter-button span {
            display: block;
            line-height: 1.1;
        }

        .filter-button span.en {
            font-size: 11px;
            color: #8292a6;
        }

        .filter-button.active,
        .filter-button:hover {
            background: var(--accent);
            color: white;
            border-color: var(--accent);
        }

        .cards-section {
            max-width: 1100px;
            margin: 0 auto;
            padding: 0 24px 64px;
        }

        .section-heading {
            font-size: 28px;
            margin: 0;
        }

        .section-heading .en {
            display: block;
            margin-top: 6px;
            font-size: 16px;
            color: #8292a6;
            letter-spacing: 0.5px;
        }

        .card-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
            gap: 24px;
            margin-top: 28px;
        }

        .card {
            background: white;
            border: 1px solid #e4ebf3;
            border-radius: 18px;
            padding: 24px;
            display: grid;
            cursor: pointer;
            gap: 14px;
            box-shadow: 0 12px 24px rgba(15, 23, 42, 0.06);
            transition: transform 0.2s ease, box-shadow 0.2s ease;
        }

        .card:hover {
            transform: translateY(-4px);
            box-shadow: 0 18px 30px rgba(15, 23, 42, 0.12);
        }

        .card-tag {
            display: inline-flex;
            flex-direction: column;
            align-items: flex-start;
            gap: 2px;
            padding: 6px 12px;
            background: var(--accent-soft);
            color: var(--accent);
            font-size: 12px;
            font-weight: 600;
            border-radius: 999px;
            text-transform: uppercase;
            letter-spacing: 1px;
        }

        .card-tag .en {
            font-size: 11px;
            color: var(--accent);
            opacity: 0.8;
        }

        .card h3 {
            margin: 0;
            font-size: 20px;
        }

        .card .title-en {
            margin: 0;
            font-size: 15px;
            color: #8292a6;
        }

        .card ul {
            margin: 0;
            padding-left: 18px;
            color: var(--text-light);
            font-size: 14px;
            display: grid;
            gap: 10px;
        }

        .card li .en {
            display: block;
            color: #8292a6;
            font-size: 12px;
            margin-top: 4px;
        }

        .insight {
            margin: 0;
            font-size: 14px;
            color: #52606d;
        }

        .insight .en {
            display: block;
            color: #9aa5b1;
            font-size: 12px;
            margin-top: 4px;
        }

        .card-actions {
            margin-top: auto;
        }

        .card-link {
            display: inline-flex;
            align-items: center;
            gap: 6px;
            padding: 10px 16px;
            background: var(--accent-soft);
            color: var(--accent);
            border-radius: 999px;
            font-size: 14px;
            font-weight: 600;
            text-decoration: none;
            transition: background 0.2s ease, transform 0.2s ease;
        }

        .card-link .en {
            font-weight: 500;
            color: var(--text-light);
        }

        .card-link:hover {
            background: rgba(12, 124, 213, 0.2);
            transform: translateY(-1px);
        }

        .models-section {
            background: var(--bg);
            padding: 72px 24px;
        }

        .models-inner {
            max-width: 1100px;
            margin: 0 auto;
            display: grid;
            gap: 32px;
        }

        .models-header h2 {
            margin: 0;
            font-size: 30px;
        }

        .models-header .en {
            display: block;
            margin-top: 6px;
            font-size: 17px;
            color: #8292a6;
        }

        .models-header p {
            margin: 12px 0 0;
            color: var(--text-light);
            max-width: 640px;
        }

        .models-header p .en {
            display: block;
            margin-top: 4px;
            color: #9aa5b1;
            font-size: 13px;
        }

        .model-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
            gap: 20px;
        }

        .model-card {
            display: grid;
            gap: 12px;
            padding: 22px;
            border-radius: 18px;
            border: 1px solid #d9e2ec;
            background: white;
            text-align: left;
            cursor: pointer;
            transition: transform 0.2s ease, box-shadow 0.2s ease, border-color 0.2s ease;
        }

        .model-card:hover {
            transform: translateY(-4px);
            border-color: var(--accent);
            box-shadow: 0 16px 28px rgba(15, 23, 42, 0.12);
        }

        .model-card h3 {
            margin: 0;
            font-size: 19px;
        }

        .model-card .title-en {
            font-size: 14px;
            color: #8292a6;
        }

        .model-card p {
            margin: 0;
            color: var(--text-light);
            font-size: 14px;
        }

        .model-card p .en {
            display: block;
            color: #9aa5b1;
            font-size: 12px;
            margin-top: 2px;
        }

        .model-badge {
            display: inline-flex;
            align-items: center;
            gap: 6px;
            font-size: 12px;
            font-weight: 600;
            letter-spacing: 1px;
            text-transform: uppercase;
            color: var(--accent);
            background: var(--accent-soft);
            border-radius: 999px;
            padding: 6px 10px;
        }

        .model-more {
            font-size: 13px;
            color: var(--accent);
            font-weight: 600;
        }

                .etf-summary-section {
            background: white;
            padding: 72px 24px;
        }

        .etf-summary-inner {
            max-width: 1100px;
            margin: 0 auto;
            display: grid;
            gap: 28px;
        }

        .etf-summary-header h2 {
            margin: 0;
            font-size: 28px;
            letter-spacing: 0.5px;
        }

        .etf-summary-header .en {
            margin-left: 8px;
            font-size: 18px;
            color: var(--text-light);
        }

        .etf-summary-header p {
            margin: 12px 0 0;
            color: var(--text-light);
            max-width: 720px;
            font-size: 15px;
        }

        .etf-summary-header p .en {
            display: inline;
        }

        .etf-card-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
            gap: 20px;
        }

        .etf-mini-card {
            background: #f0f7ff;
            border: 1px solid #d2e3f8;
            border-radius: 18px;
            padding: 24px;
            display: grid;
            gap: 14px;
            text-align: left;
            cursor: pointer;
            transition: transform 0.2s ease, box-shadow 0.2s ease, border-color 0.2s ease;
        }

        .etf-mini-card:hover {
            transform: translateY(-4px);
            border-color: var(--accent);
            box-shadow: 0 16px 28px rgba(12, 124, 213, 0.18);
        }

        .etf-mini-card h3 {
            margin: 0;
            font-size: 20px;
        }

        .etf-mini-card h3 .en {
            display: block;
            margin-top: 4px;
            font-size: 14px;
            color: #4d5d6c;
        }

        .etf-mini-card p {
            margin: 0;
            color: var(--text-light);
            font-size: 14px;
        }

        .etf-mini-card p .en {
            display: block;
            margin-top: 4px;
            color: #9aa5b1;
            font-size: 12px;
        }

        .etf-mini-card .meta {
            display: flex;
            gap: 16px;
            font-size: 13px;
            color: #4f5d75;
        }

        .etf-mini-card .meta strong {
            font-size: 18px;
            color: var(--accent);
        }

        .etf-mini-card .meta .en {
            display: block;
            font-size: 11px;
            color: #8292a6;
        }

        .etf-overlay {
            position: fixed;
            inset: 0;
            background: rgba(15, 23, 42, 0.5);
            display: none;
            align-items: center;
            justify-content: center;
            padding: 32px;
            z-index: 15;
        }

        .etf-overlay.active {
            display: flex;
        }

        .etf-detail {
            background: white;
            border-radius: 22px;
            max-width: 1000px;
            width: 100%;
            padding: 32px;
            display: grid;
            gap: 24px;
            box-shadow: 0 32px 56px rgba(15, 23, 42, 0.35);
            position: relative;
        }

        .etf-detail-header h3 {
            margin: 0;
            font-size: 24px;
        }

        .etf-detail-header h3 .en {
            display: block;
            margin-top: 6px;
            font-size: 15px;
            color: #8292a6;
        }

        .etf-detail-header p {
            margin: 8px 0 0;
            color: var(--text-light);
            font-size: 14px;
        }

        .etf-detail-layout {
            display: grid;
            grid-template-columns: 320px 1fr;
            gap: 24px;
        }

        .etf-search {
            position: relative;
        }

        .etf-search input {
            width: 100%;
            padding: 10px 14px;
            border-radius: 12px;
            border: 1px solid #d9e2ec;
            font-size: 14px;
        }

        .etf-list {
            list-style: none;
            margin: 16px 0 0;
            padding: 0;
            border: 1px solid #e4ebf3;
            border-radius: 16px;
            max-height: 420px;
            overflow: auto;
        }

        .etf-list-item {
            display: grid;
            gap: 6px;
            padding: 14px 16px;
            border-bottom: 1px solid #e4ebf3;
            cursor: pointer;
            background: white;
            transition: background 0.2s ease;
        }

        .etf-list-item:last-child {
            border-bottom: none;
        }

        .etf-list-item:hover,
        .etf-list-item.active {
            background: rgba(12, 124, 213, 0.08);
        }

        .etf-list-item strong {
            font-size: 15px;
        }

        .etf-list-item .ticker {
            font-size: 12px;
            color: var(--accent);
            text-transform: uppercase;
        }

        .etf-list-item .stats {
            display: flex;
            gap: 12px;
            font-size: 12px;
            color: #4f5d75;
        }

        .etf-chart-panel {
            background: #f8fbff;
            border-radius: 18px;
            padding: 20px;
            border: 1px solid #d9e2ec;
            display: grid;
            gap: 18px;
            min-height: 460px;
        }

        .etf-chart-panel canvas {
            width: 100%;
            max-height: 320px;
        }

        .etf-chart-empty {
            display: grid;
            place-items: center;
            color: #9aa5b1;
            font-size: 13px;
            height: 320px;
        }

        .etf-chart-meta {
            display: grid;
            gap: 10px;
            font-size: 13px;
            color: #4f5d75;
        }

        .etf-chart-meta .row {
            display: flex;
            justify-content: space-between;
            align-items: center;
        }

        .etf-chart-meta .label {
            font-weight: 600;
            color: var(--text-dark);
        }

        .etf-chart-meta .value {
            font-variant-numeric: tabular-nums;
        }

        .etf-close {
            position: absolute;
            top: 18px;
            right: 18px;
            background: none;
            border: none;
            font-size: 26px;
            cursor: pointer;
            color: #708090;
        }

        @media (max-width: 960px) {
            .etf-detail-layout {
                grid-template-columns: 1fr;
            }

            .etf-chart-panel {
                min-height: 360px;
            }
        }

        @media (max-width: 640px) {
            .etf-summary-section {
                padding: 48px 16px;
            }

            .etf-card-grid {
                grid-template-columns: 1fr;
            }

            .etf-detail {
                padding: 22px;
            }

            .etf-overlay {
                padding: 16px;
            }
        }
        .model-overlay {
            position: fixed;
            inset: 0;
            background: rgba(15, 23, 42, 0.45);
            display: none;
            align-items: center;
            justify-content: center;
            padding: 24px;
            z-index: 10;
        }

        .model-overlay.active {
            display: flex;
        }

        .model-detail {
            background: white;
            border-radius: 20px;
            max-width: 560px;
            width: 100%;
            padding: 32px;
            display: grid;
            gap: 18px;
            position: relative;
            box-shadow: 0 28px 48px rgba(15, 23, 42, 0.28);
        }

        .model-detail h3 {
            margin: 0;
        }

        .model-detail h3 .en {
            display: block;
            color: #8292a6;
            font-size: 15px;
            margin-top: 4px;
        }

        .model-detail p {
            margin: 0;
            color: var(--text-light);
        }

        .model-detail p .en {
            display: block;
            color: #9aa5b1;
            font-size: 13px;
            margin-top: 4px;
        }

        .model-detail ul {
            margin: 0;
            padding-left: 18px;
            color: var(--text-light);
            font-size: 14px;
            display: grid;
            gap: 10px;
        }

        .model-detail li .en {
            display: block;
            color: #9aa5b1;
            font-size: 12px;
            margin-top: 4px;
        }

        .close-detail {
            position: absolute;
            top: 16px;
            right: 16px;
            background: none;
            border: none;
            font-size: 24px;
            cursor: pointer;
            color: #8292a6;
        }

        .card-overlay {
            position: fixed;
            inset: 0;
            background: rgba(15, 23, 42, 0.55);
            display: none;
            align-items: center;
            justify-content: center;
            padding: 24px;
            z-index: 12;
        }

        .card-overlay.active {
            display: flex;
        }

        .card-overlay-panel {
            background: white;
            border-radius: 18px;
            max-width: 520px;
            width: 100%;
            padding: 28px;
            display: grid;
            gap: 18px;
            position: relative;
            box-shadow: 0 24px 48px rgba(15, 23, 42, 0.3);
        }

        .card-overlay-panel h3 {
            margin: 0;
        }

        .card-overlay-panel h3 .en {
            display: block;
            color: #8292a6;
            font-size: 14px;
            margin-top: 4px;
        }

        .card-overlay-panel p {
            margin: 0;
            color: var(--text-light);
            font-size: 15px;
        }

        .card-overlay-panel p .en {
            display: block;
            color: #9aa5b1;
            font-size: 13px;
            margin-top: 6px;
        }

        .card-overlay-actions {
            display: flex;
            flex-wrap: wrap;
            gap: 12px;
            justify-content: flex-end;
            margin-top: 8px;
        }

        .card-overlay-actions .ghost {
            background: transparent;
            border: 1px solid #d5dee9;
            color: var(--text-light);
            padding: 10px 18px;
            border-radius: 999px;
            cursor: pointer;
        }

        .card-overlay-actions .primary {
            background: var(--accent);
            color: white;
            padding: 10px 20px;
            border-radius: 999px;
            text-decoration: none;
            display: inline-flex;
            align-items: center;
            gap: 8px;
            font-weight: 600;
        }

        .card-overlay-close {
            position: absolute;
            top: 16px;
            right: 16px;
            background: none;
            border: none;
            font-size: 24px;
            cursor: pointer;
            color: #8a99ac;
        }

        .card-detail-page {
            max-width: 1100px;
            margin: 0 auto;
            padding: 72px 24px 80px;
            display: grid;
            gap: 32px;
        }

        .card-detail-header {
            display: grid;
            gap: 12px;
        }

        .card-detail-header .card-tag {
            display: inline-flex;
            align-items: center;
            gap: 8px;
            font-size: 14px;
            color: var(--accent);
            background: var(--accent-soft);
            padding: 6px 12px;
            border-radius: 999px;
            width: fit-content;
        }

        .card-detail-header h1 {
            margin: 0;
            font-size: clamp(32px, 5vw, 40px);
            line-height: 1.2;
        }

        .card-detail-header h1 .en {
            display: block;
            color: #8292a6;
            font-size: 16px;
            margin-top: 6px;
        }

        .card-detail-overview {
            font-size: 17px;
            color: var(--text-light);
        }

        .card-detail-overview .en {
            display: block;
            color: #9aa5b1;
            font-size: 14px;
            margin-top: 6px;
        }

        .card-detail-body {
            display: grid;
            gap: 28px;
        }

        .card-detail-section h2 {
            margin: 0 0 12px;
            font-size: 20px;
        }

        .card-detail-section h2 .en {
            display: block;
            color: #8292a6;
            font-size: 14px;
            margin-top: 4px;
        }

        .card-detail-section ul {
            margin: 0;
            padding-left: 20px;
            display: grid;
            gap: 12px;
            color: var(--text-light);
        }

        .card-detail-section li .en {
            display: block;
            color: #9aa5b1;
            font-size: 13px;
            margin-top: 4px;
        }

        .card-detail-related {
            border-top: 1px solid #e4ebf3;
            padding-top: 32px;
            display: grid;
            gap: 16px;
        }

        .card-detail-related-grid {
            display: grid;
            gap: 16px;
            grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
        }

        .card-detail-related a {
            border: 1px solid #e4ebf3;
            border-radius: 14px;
            padding: 18px;
            text-decoration: none;
            color: inherit;
            background: white;
            transition: border-color 0.2s ease, box-shadow 0.2s ease;
        }

        .card-detail-related a:hover {
            border-color: var(--accent);
            box-shadow: 0 12px 24px rgba(12, 124, 213, 0.12);
        }

        @media (max-width: 640px) {
            .card-overlay-panel {
                padding: 22px;
            }

            .card-detail-page {
                padding: 56px 16px 64px;
            }
        }

        footer {
            padding: 24px;
            background: white;
            border-top: 1px solid #e4ebf3;
            color: var(--text-light);
            font-size: 13px;
        }

        footer .footer-content {
            max-width: 1100px;
            margin: 0 auto;
            display: flex;
            flex-wrap: wrap;
            gap: 12px;
            justify-content: space-between;
            align-items: center;
        }

        footer ul {
            display: flex;
            gap: 16px;
            margin: 0;
            padding: 0;
            list-style: none;
        }

        footer ul li span.en {
            display: block;
            font-size: 11px;
            color: #9aa5b1;
        }

        @media (max-width: 640px) {
            nav {
                display: none;
            }

            .top-bar {
                padding: 16px;
            }

            .model-grid {
                grid-template-columns: 1fr;
            }
        }
    </style>
</head>
<body data-page="{{ page_type }}">
    <header>
        <div class='top-bar'>
            <div class='brand'>
                <div class='brand-icon'>YJ</div>
                yjc
            </div>
            <nav>
                {% if page_type == 'home' %}
                    {% for category in categories if category.id != 'all' %}
                    <a href='#{{ 'models' if category.id == 'macro' else 'cards' }}'>{{ category.label_zh }} / {{ category.label_en }}</a>
                    {% endfor %}
                    <a href='#etfs'>ETF / ETF</a>
                {% else %}
                    <a href="/" class="{{ 'active' if page_type == 'card_detail' else '' }}">返回主页 / Home</a>
                    <a href="/#cards" class="{{ 'active' if page_type == 'card_detail' else '' }}">知识卡片 / Cards</a>
                    <a href="/#models">模型精选 / Models</a>
                    <a href="/#etfs" class="{{ 'active' if page_type == 'etf-detail' else '' }}">ETF 专区 / ETF</a>
                {% endif %}
                <a href='#footer'>Contact</a>
            </nav>
            <button class='cta' type='button'>快速速查 / Quick Reference</button>
        </div>
    </header>

    {% if page_type == 'home' %}
    <section class='hero'>
        <div>
            <h1>轻松掌握金融常识，双语速读核心要点</h1>
            <span class='headline-en'>Master finance fundamentals faster with bilingual flashcards.</span>
        </div>
        <p>
            从基础概念到实务模型，精选投资、理财、风险控制与宏观洞察，帮助你迅速搭建知识框架。
            <span class='para-en'>From fundamentals to hands-on models, explore investing, banking, risk, and macro insights side by side.</span>
        </p>
        <div class='filter-bar' id='filterBar'>
            <span class='filter-title'>热门主题 / Topics</span>
            {% for category in categories %}
            <button class='filter-button {% if category.id == "all" %}active{% endif %}' data-filter='{{ category.id }}'>
                <span>{{ category.label_zh }}</span>
                <span class='en'>{{ category.label_en }}</span>
            </button>
            {% endfor %}
        </div>
    </section>

    <section class='cards-section' id='cards'>
        <h2 class='section-heading'>知识卡片精选<span class='en'>Essential Finance Knowledge Cards</span></h2>
        <div class='card-grid'>
            {% for card in cards %}
            <article class='card' data-category='{{ card.category }}' data-slug='{{ card.slug }}'>
                <span class='card-tag'>
                    {{ card.tag_zh }}
                    <span class='en'>{{ card.tag_en }}</span>
                </span>
                <h3>{{ card.title_zh }}</h3>
                <p class='title-en'>{{ card.title_en }}</p>
                <ul>
                    {% for bullet in card.bullets %}
                    <li>
                        {{ bullet.zh }}
                        <span class='en'>{{ bullet.en }}</span>
                    </li>
                    {% endfor %}
                </ul>
                <p class='insight'>
                    {{ card.insight_zh }}
                    <span class='en'>{{ card.insight_en }}</span>
                </p>
                <div class='card-actions'>
                    <a class='card-link' href='{{ card.detail_url }}'>
                        查看卡片详情<span class='en'>Open detailed insight</span>
                    </a>
                </div>
            </article>
            {% endfor %}
        </div>
    </section>
    <div class='card-overlay' id='cardOverlay'>
        <div class='card-overlay-panel' role='dialog' aria-modal='true'>
            <button class='card-overlay-close' type='button' id='cardOverlayClose' aria-label='关闭卡片弹窗'>&times;</button>
            <h3 id='cardOverlayTitle'></h3>
            <p id='cardOverlayOverview'></p>
            <div class='card-overlay-actions'>
                <button class='ghost' type='button' id='cardOverlayCancel'>继续浏览</button>
                <a class='primary' id='cardOverlayGo' href='#'>前往详情<span class='en'>Go to detail</span></a>
            </div>
        </div>
    </div>

    <section class='models-section' id='models'>
        <div class='models-inner'>
            <div class='models-header'>
                <h2>金融模型小卡片<span class='en'>Finance Model Spotlights</span></h2>
                <p>
                    点击了解常用模型的核心公式、适用场景与实操要点。
                    <span class='en'>Tap a card to explore formulas, use cases, and implementation tips in seconds.</span>
                </p>
            </div>
            <div class='model-grid'>
                {% for model in models %}
                <button class='model-card' type='button' data-model='{{ model.id }}'>
                    <span class='model-badge'>模型 / Model</span>
                    <h3>{{ model.title_zh }}<span class='title-en'>{{ model.title_en }}</span></h3>
                    <p>
                        {{ model.subtitle_zh }}
                        <span class='en'>{{ model.subtitle_en }}</span>
                    </p>
                    <span class='model-more'>点击查看详情 → / Explore Details →</span>
                </button>
                {% endfor %}
            </div>
        </div>
    </section>

    <section class='etf-summary-section' id='etfs'>
        <div class='etf-summary-inner'>
            <div class='etf-summary-header'>
                <h2>ETF 热门导航<span class='en'>ETF Discovery</span></h2>
                <p>
                    通过一张小卡片快速洞察热门ETF，再展开查看20只基金的核心指标与实时收益曲线。
                    <span class='en'>Start from spotlight cards, explore 30+ tracked ETFs, and unlock fresh performance insights.</span>
                </p>
            </div>
            <div class='etf-card-grid'>
                {% for card in etf_cards %}
                <a class='etf-mini-card' href='/etf/{{ card.id }}'>
                    <h3>{{ card.title_zh }}<span class='en'>{{ card.title_en }}</span></h3>
                    <p>
                        {{ card.description_zh }}
                        <span class='en'>{{ card.description_en }}</span>
                    </p>
                    <div class='meta'>
                        <span>
                            <strong>{{ etfs|length }}</strong>
                            <span class='en'>Fund count</span>
                        </span>
                        <span>
                            {% if card.asset_filters %}
                            {{ card.asset_filters|length }} 类主题<span class='en'>Focus styles</span>
                            {% else %}
                            全市场覆盖<span class='en'>All segments</span>
                            {% endif %}
                        </span>
                    </div>
                </button>
                {% endfor %}
            </div>
        </div>
    </section>

    <div class='model-overlay' id='modelOverlay'>
        <div class='model-detail'>
            <button class='close-detail' type='button' id='closeModelDetail'>&times;</button>
            <h3 id='modelDetailTitle'></h3>
            <p id='modelDetailDescription'></p>
            <ul id='modelDetailHighlights'></ul>
        </div>
    </div>
    {% elif page_type == 'card_detail' %}
    <main class='card-detail-page'>
        <div class='card-detail-header'>
            <span class='card-tag'>
                {{ card_detail.card.tag_zh }}
                <span class='en'>{{ card_detail.card.tag_en }}</span>
            </span>
            <h1>{{ card_detail.card.title_zh }}<span class='en'>{{ card_detail.card.title_en }}</span></h1>
            <div class='card-detail-overview'>
                {{ card_detail.overview_zh }}
                <span class='en'>{{ card_detail.overview_en }}</span>
            </div>
        </div>
        <div class='card-detail-body'>
            <section class='card-detail-section'>
                <h2>要点展开<span class='en'>Deep Dive</span></h2>
                <ul>
                    {% for point in card_detail.deep_dives %}
                    <li>
                        {{ point.zh }}
                        <span class='en'>{{ point.en }}</span>
                    </li>
                    {% endfor %}
                </ul>
            </section>
            <section class='card-detail-section'>
                <h2>核心提示<span class='en'>Insight</span></h2>
                <p>
                    {{ card_detail.card.insight_zh }}
                    <span class='en'>{{ card_detail.card.insight_en }}</span>
                </p>
            </section>
        </div>
        {% if related_cards %}
        <div class='card-detail-related'>
            <h2>同类卡片推荐<span class='en'>You may also explore</span></h2>
            <div class='card-detail-related-grid'>
                {% for related in related_cards %}
                <a href='{{ related.detail_url }}'>
                    <strong>{{ related.title_zh }}</strong>
                    <span class='en'>{{ related.title_en }}</span>
                </a>
                {% endfor %}
            </div>
        </div>
        {% endif %}
    </main>
{% elif page_type == 'etf-detail' %}
    <section class='detail-hero'>
        <div class='detail-hero-text'>
            <h1>{{ card.title_zh }}<span class='headline-en'>{{ card.title_en }}</span></h1>
            <p>
                {{ card.description_zh }}
                <span class='para-en'>{{ card.description_en }}</span>
            </p>
            {% if card.asset_filters %}
            <div class='detail-tags'>
                {% for tag in card.asset_filters %}
                <span class='detail-tag'>{{ tag }}</span>
                {% endfor %}
            </div>
            {% endif %}
        </div>
        <div class='detail-hero-meta'>
            <a class='cta secondary' href='/#etfs'>?? ETF ??<span class='en'>Back to overview</span></a>
        </div>
    </section>

    <section class='etf-detail-section'>
        <div class='etf-detail-layout'>
            <aside class='etf-detail-sidebar'>
                <div class='sidebar-header'>
                    <h2>????<span class='en'>Fund Lineup</span></h2>
                    <input type='search' id='detailEtfSearch' placeholder='????????? / Filter by name or ticker'>
                </div>
                <ul class='detail-etf-list' id='detailEtfList'></ul>
                <div class='sidebar-note'>
                    ?????????????????<span class='en'>Window: latest ~35 trading days.</span>
                </div>
            </aside>
            <div class='etf-detail-main'>
                <div class='chart-headline'>
                    {% if matched_etfs %}
                    <h2 id='detailEtfTitle'>{{ matched_etfs[0].name }}<span class='en'>{{ matched_etfs[0].ticker }}</span></h2>
                    <p id='detailEtfSubtitle'>{{ matched_etfs[0].asset_class }}</p>
                    {% else %}
                    <h2 id='detailEtfTitle'>????</h2>
                    <p id='detailEtfSubtitle'></p>
                    {% endif %}
                </div>
                <div class='detail-chart-panel'>
                    <canvas id='detailEtfChart' aria-label='ETF performance chart'></canvas>
                    <div class='etf-chart-meta' id='detailEtfMeta'></div>
                </div>
            </div>
        </div>
    </section>

    <section class='detail-table-section'>
        <div class='detail-table-inner'>
            <h2>ETF ????<span class='en'>Snapshot Metrics</span></h2>
            <table class='etf-table'>
                <thead>
                    <tr>
                        <th>?? / ??<span class='en'>Fund / Ticker</span></th>
                        <th>??<span class='en'>Category</span></th>
                        <th>???<span class='en'>Provider</span></th>
                        <th>??<span class='en'>AUM</span></th>
                        <th>??<span class='en'>Expense</span></th>
                        <th>????<span class='en'>YTD</span></th>
                        <th>???<span class='en'>1Y</span></th>
                    </tr>
                </thead>
                <tbody id='detailEtfTable'>
                    {% if not matched_etfs %}
                    <tr><td colspan='7'>??????</td></tr>
                    {% endif %}
                </tbody>
            </table>
        </div>
    </section>

    <section class='related-section'>
        <div class='related-inner'>
            <h2>?? yjc ??<span class='en'>More from yjc</span></h2>
            <div class='related-grid'>
                <a class='related-card' href='/#cards'>
                    <h3>??????<span class='en'>Knowledge Cards</span></h3>
                    <p>??????????????????</p>
                </a>
                <a class='related-card' href='/#models'>
                    <h3>??????<span class='en'>Model Spotlights</span></h3>
                    <p>??????????????????</p>
                </a>
                <a class='related-card' href='/#etfs'>
                    <h3>?? ETF ??<span class='en'>Back to ETF Discovery</span></h3>
                    <p>??????????????</p>
                </a>
            </div>
        </div>
    </section>
    {% endif %}

    <footer id='footer'>
        <div class='footer-content'>
            <span>? 2025 yjc · 以知识赋能投资者 / Empowering investors through knowledge.</span>
            <ul>
                <li><a href='#'>使用条款<span class='en'>Terms</span></a></li>
                <li><a href='#'>隐私政策<span class='en'>Privacy</span></a></li>
                <li><a href='#'>联系我们<span class='en'>Contact</span></a></li>
            </ul>
        </div>
    </footer>

    <script>
        const pageType = document.body.dataset.page || 'home';

        if (pageType === "home") {
            const filterButtons = document.querySelectorAll('.filter-button');
            const cards = document.querySelectorAll('.card');

            filterButtons.forEach(button => {
                button.addEventListener('click', () => {
                    const category = button.dataset.filter;

                    filterButtons.forEach(btn => btn.classList.remove('active'));
                    button.classList.add('active');

                    cards.forEach(card => {
                        const match = category === 'all' || card.dataset.category === category;
                        card.style.display = match ? 'grid' : 'none';
                    });
                });
            });

            const cardsData = {{ cards|tojson }};
            const cardOverlay = document.getElementById('cardOverlay');
            const cardOverlayTitle = document.getElementById('cardOverlayTitle');
            const cardOverlayOverview = document.getElementById('cardOverlayOverview');
            const cardOverlayClose = document.getElementById('cardOverlayClose');
            const cardOverlayCancel = document.getElementById('cardOverlayCancel');
            const cardOverlayGo = document.getElementById('cardOverlayGo');
            const cardLookup = {};
            cardsData.forEach(item => {
                cardLookup[item.slug] = item;
            });

            const hideCardOverlay = () => {
                if (cardOverlay) {
                    cardOverlay.classList.remove('active');
                }
            };

            cards.forEach(card => {
                card.addEventListener('click', event => {
                    if (event.target.closest('.card-link')) {
                        return;
                    }
                    const slug = card.dataset.slug;
                    const data = cardLookup[slug];

                    if (!cardOverlay || !cardOverlayTitle || !cardOverlayOverview || !cardOverlayGo || !data) {
                        window.location.href = `/card/${slug}`;
                        return;
                    }

                    const overviewZh = data.overview_zh || data.insight_zh || '';
                    const overviewEn = data.overview_en || data.insight_en || '';

                    cardOverlayTitle.innerHTML = `${data.title_zh}<span class="en">${data.title_en}</span>`;
                    cardOverlayOverview.innerHTML = `${overviewZh}<span class="en">${overviewEn}</span>`;
                    cardOverlayGo.href = data.detail_url || `/card/${slug}`;
                    cardOverlay.classList.add('active');
                });
            });

            if (cardOverlayClose) {
                cardOverlayClose.addEventListener('click', hideCardOverlay);
            }
            if (cardOverlayCancel) {
                cardOverlayCancel.addEventListener('click', event => {
                    event.preventDefault();
                    hideCardOverlay();
                });
            }
            if (cardOverlay) {
                cardOverlay.addEventListener('click', event => {
                    if (event.target === cardOverlay) {
                        hideCardOverlay();
                    }
                });
            }
            document.addEventListener('keydown', event => {
                if (event.key === 'Escape' && cardOverlay && cardOverlay.classList.contains('active')) {
                    hideCardOverlay();
                }
            });

            const modelsData = {{ models|tojson }};
            const modelOverlay = document.getElementById('modelOverlay');
            const modelCards = document.querySelectorAll('.model-card');
            const modelDetailTitle = document.getElementById('modelDetailTitle');
            const modelDetailDescription = document.getElementById('modelDetailDescription');
            const modelDetailHighlights = document.getElementById('modelDetailHighlights');
            const closeModelDetail = document.getElementById('closeModelDetail');

            modelCards.forEach(card => {
                card.addEventListener('click', () => {
                    const modelId = card.dataset.model;
                    const model = modelsData.find(item => item.id === modelId);
                    if (!model) {
                        return;
                    }

                    modelDetailTitle.innerHTML = `${model.title_zh}<span class="en">${model.title_en}</span>`;
                    modelDetailDescription.innerHTML = `${model.description_zh}<span class="en">${model.description_en}</span>`;
                    modelDetailHighlights.innerHTML = '';
                    model.highlights.forEach(point => {
                        const li = document.createElement('li');
                        li.innerHTML = `${point.zh}<span class="en">${point.en}</span>`;
                        modelDetailHighlights.appendChild(li);
                    });

                    modelOverlay.classList.add('active');
                });
            });

            if (closeModelDetail) {
                closeModelDetail.addEventListener('click', () => {
                    modelOverlay.classList.remove('active');
                });
            }

            if (modelOverlay) {
                modelOverlay.addEventListener('click', event => {
                    if (event.target === modelOverlay) {
                        modelOverlay.classList.remove('active');
                    }
                });
            }
        }
    </script>

</body>
</html>"""

def _format_series(raw_pairs):
    if not raw_pairs:
        return []

    ordered = sorted(raw_pairs, key=lambda item: item[0])
    if len(ordered) > 35:
        ordered = ordered[-35:]
    base = ordered[0][1] or 1
    if base == 0:
        base = 1

    series = []
    for date_str, value in ordered:
        return_pct = round((value / base - 1) * 100, 2)
        series.append({
            'date': date_str,
            'nav': round(value, 4),
            'return_pct': return_pct,
        })
    return series

def _request_yahoo_series(symbol: str):
    symbol = (symbol or '').upper()
    if not symbol:
        return []

    url = f'https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?interval=1d&range=6mo'
    try:
        response = requests.get(url, timeout=6)
        response.raise_for_status()
        payload = response.json()
    except Exception:
        return []

    result = (payload.get('chart') or {}).get('result')
    if not result:
        return []

    first = result[0]
    timestamps = first.get('timestamp') or []
    quote = (first.get('indicators') or {}).get('quote') or [{}]
    closes = quote[0].get('close') or []

    raw_pairs = []
    for ts, close in zip(timestamps, closes):
        if close in (None, 'null'):
            continue
        try:
            date_str = datetime.fromtimestamp(ts).strftime('%Y-%m-%d')
        except Exception:
            continue
        raw_pairs.append((date_str, float(close)))

    return _format_series(raw_pairs)

def _request_eastmoney_series(code: str):
    code = (code or '').strip()
    if not code:
        return []

    params = {
        'FCODE': code,
        'pageIndex': 1,
        'pageSize': 35,
        'appType': 'ttjj',
        'product': 'EFund',
        'plat': 'Iphone',
        'deviceid': '00000000-0000-0000-0000-000000000000',
        'Version': '1',
        'lf': '1',
        'ctoken': '',
        'userId': '',
        'u': '0',
        'UToken': '',
        'range': '1m',
    }

    try:
        response = requests.get(
            EASTMONEY_HISTORY_URL,
            params=params,
            headers=EASTMONEY_HEADERS,
            timeout=6,
        )
        response.raise_for_status()
        payload = response.json()
    except Exception:
        return []

    items = ((payload.get('Data') or {}).get('LSJZList')) or (payload.get('Datas') or [])
    raw_pairs = []
    for item in items:
        date_str = item.get('FSRQ')
        value_str = item.get('DWJZ')
        if not date_str or value_str in (None, ''):
            continue
        try:
            value = float(value_str)
        except (TypeError, ValueError):
            continue
        raw_pairs.append((date_str, value))

    return _format_series(raw_pairs)


def _write_records_to_excel(records):
    wb = Workbook()
    ws = wb.active
    ws.title = 'etf_history'
    ws.append(['ticker', 'date', 'nav', 'return_pct'])
    for record in records:
        ws.append([
            record['ticker'],
            record['date'],
            record['nav'],
            record['return_pct'],
        ])

    meta = wb.create_sheet('meta')
    meta.append(['generated_at', datetime.utcnow().isoformat(timespec='seconds')])
    wb.save(EXCEL_PATH)


def _load_cache_from_excel():
    if not EXCEL_PATH.exists():
        return {}

    try:
        wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except Exception:
        return {}

    if 'etf_history' not in wb.sheetnames:
        return {}

    sheet = wb['etf_history']
    cache = {}
    for ticker, date_value, nav_value, return_value in sheet.iter_rows(min_row=2, values_only=True):
        if not ticker or not date_value:
            continue
        if isinstance(date_value, datetime):
            date_str = date_value.strftime('%Y-%m-%d')
        else:
            date_str = str(date_value)
        try:
            nav = float(nav_value)
        except (TypeError, ValueError):
            continue
        try:
            pct = float(return_value)
        except (TypeError, ValueError):
            pct = 0.0
        cache.setdefault(ticker.upper(), []).append({
            'date': date_str,
            'nav': round(nav, 4),
            'return_pct': round(pct, 2),
        })

    for series in cache.values():
        series.sort(key=lambda item: item['date'])
    return cache


def _group_records(records):
    grouped = {}
    for record in records:
        grouped.setdefault(record['ticker'], []).append(record)
    for series in grouped.values():
        series.sort(key=lambda item: item['date'])
    return grouped


def _refresh_cache_from_remote():
    aggregated = []
    for etf in ETFS:
        ticker = etf['ticker']
        series = _fetch_remote_etf_series(ticker)
        if not series:
            continue
        for point in series:
            aggregated.append({
                'ticker': ticker,
                'date': point['date'],
                'nav': point['nav'],
                'return_pct': point['return_pct'],
            })

    if not aggregated:
        return False

    _write_records_to_excel(aggregated)
    global ETF_CACHE, ETF_CACHE_MTIME
    ETF_CACHE = _group_records(aggregated)
    ETF_CACHE_MTIME = time.time()
    return True


def ensure_etf_cache(force_refresh=False):
    global ETF_CACHE, ETF_CACHE_MTIME

    if force_refresh:
        if _refresh_cache_from_remote():
            return True

    now = time.time()

    if ETF_CACHE and (now - ETF_CACHE_MTIME) < CACHE_MAX_AGE_SECONDS:
        return True

    if EXCEL_PATH.exists():
        mtime = EXCEL_PATH.stat().st_mtime
        if not ETF_CACHE or ETF_CACHE_MTIME != mtime:
            cache = _load_cache_from_excel()
            if cache:
                ETF_CACHE = cache
                ETF_CACHE_MTIME = mtime
        if ETF_CACHE and (now - mtime) < CACHE_MAX_AGE_SECONDS:
            return True

    if _refresh_cache_from_remote():
        return True

    if ETF_CACHE:
        return True

    if EXCEL_PATH.exists():
        cache = _load_cache_from_excel()
        if cache:
            ETF_CACHE = cache
            ETF_CACHE_MTIME = EXCEL_PATH.stat().st_mtime
            return True

    return False


def fetch_etf_series(ticker: str, force_refresh=False):
    normalized = (ticker or '').upper()
    if not normalized:
        return []

    if not ensure_etf_cache(force_refresh=force_refresh):
        return []

    return ETF_CACHE.get(normalized, [])

def _fetch_remote_etf_series(ticker: str):
    base = (ticker or '').upper()
    if not base:
        return []

    series = _request_eastmoney_series(base)
    if series:
        return series

    candidates = []
    mapped = ETF_SYMBOL_MAP.get(base)
    if mapped:
        candidates.append(mapped.upper())

    for suffix in ['', '.SS', '.SZ']:
        candidate = base if suffix == '' else f"{base}{suffix}"
        if candidate not in candidates:
            candidates.append(candidate)

    for symbol in candidates:
        series = _request_yahoo_series(symbol)
        if series:
            return series

    return []

